from __future__ import annotations

import argparse
from collections.abc import Iterable, Sequence
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font

from odoo_ai_manager.config import OdooSettings
from odoo_ai_manager.domain.ports import OdooReadClient
from odoo_ai_manager.domain.models import AccessMode
from odoo_ai_manager.infrastructure.odoo.xmlrpc import XmlRpcOdooClient
from odoo_ai_manager.paths import project_root

from .entities import DailySalesReportConfig, PosPayment, PosSale


_ORDER_STATES = ["paid", "done", "invoiced"]
_PAGE_SIZE = 1000


def generate_daily_sales_report(
    client: OdooReadClient,
    config: DailySalesReportConfig,
) -> Path:
    """Consulta ventas y crea un XLSX en el timezone del usuario de Odoo."""

    domain: list[tuple[str, str, Any]] = [
        (
            "order_id.date_order",
            ">=",
            _odoo_datetime(
                datetime.combine(
                    config.start_date,
                    time.min,
                    tzinfo=client.user_timezone,
                )
            ),
        ),
        ("order_id.state", "in", _ORDER_STATES),
    ]
    company_id = getattr(client, "company_id", None)
    if company_id is not None:
        domain.append(("order_id.company_id", "=", company_id))
    if config.end_date is not None:
        exclusive_end = config.end_date + timedelta(days=1)
        domain.append(
            (
                "order_id.date_order",
                "<",
                _odoo_datetime(
                    datetime.combine(
                        exclusive_end,
                        time.min,
                        tzinfo=client.user_timezone,
                    )
                ),
            )
        )

    lines = _search_read_all(
        client,
        "pos.order.line",
        domain,
        ["id", "order_id", "product_id", "qty", "price_unit"],
        order="order_id, id",
    )
    sales = _build_sales(client, lines)
    return DailySalesXlsxWriter().save(sales, config.output_path)


def _build_sales(
    client: OdooReadClient,
    lines: Sequence[dict[str, Any]],
) -> list[PosSale]:
    if not lines:
        return []

    order_ids = _relation_ids(line.get("order_id") for line in lines)
    orders = _search_read_all(
        client,
        "pos.order",
        [("id", "in", order_ids)],
        [
            "id",
            "name",
            "pos_reference",
            "date_order",
            "config_id",
            "session_id",
            "payment_ids",
        ],
    )
    order_by_id = {
        int(order["id"]): order
        for order in orders
        if order.get("id") is not None
    }

    product_ids = _relation_ids(line.get("product_id") for line in lines)
    products = _search_read_all(
        client,
        "product.product",
        [("id", "in", product_ids)],
        ["id", "default_code"],
    )
    product_reference_by_id = {
        int(product["id"]): product.get("default_code") or None
        for product in products
        if product.get("id") is not None
    }

    payment_ids = _list_relation_ids(
        order.get("payment_ids") for order in orders
    )
    payments = _search_read_all(
        client,
        "pos.payment",
        [("id", "in", payment_ids)],
        ["id", "payment_method_id", "amount"],
    )
    payment_by_id = {
        int(payment["id"]): PosPayment(
            payment_id=int(payment["id"]),
            payment_type=_relation_name(payment.get("payment_method_id"))
            or "Sin pago",
            amount=_decimal_value(payment.get("amount")),
        )
        for payment in payments
        if payment.get("id") is not None
    }
    payments_by_order = {
        int(order["id"]): _payments_for_order(
            order.get("payment_ids", []),
            payment_by_id,
        )
        for order in orders
        if order.get("id") is not None
    }

    sales: list[PosSale] = []
    for line in lines:
        order_id = _relation_id(line.get("order_id"))
        product_id = _relation_id(line.get("product_id"))
        if order_id is None or product_id is None:
            continue
        order = order_by_id.get(order_id)
        if order is None:
            continue
        order_payments = payments_by_order.get(order_id, [])
        sales.append(
            PosSale(
                date=_parse_odoo_datetime(
                    order.get("date_order"),
                    client.user_timezone,
                ),
                payment_type=_payment_type_label(order_payments),
                product=_relation_name(line.get("product_id")) or "Sin nombre",
                product_reference=product_reference_by_id.get(product_id),
                pos_name=_relation_name(order.get("config_id")) or "Sin POS",
                sale_reference=(
                    order.get("pos_reference")
                    or order.get("name")
                    or "Sin referencia"
                ),
                pos_session=(
                    _relation_name(order.get("session_id")) or "Sin sesion"
                ),
                payments=order_payments,
                quantity=_decimal_value(line.get("qty")),
                price=_decimal_value(line.get("price_unit")),
            )
        )
    return sales


def _search_read_all(
    client: OdooReadClient,
    model: str,
    domain: Sequence[Sequence[Any]],
    fields: Sequence[str],
    order: str | None = None,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    offset = 0
    while True:
        page = client.search_read(
            model,
            domain,
            fields,
            offset=offset,
            limit=_PAGE_SIZE,
            order=order,
        )
        records.extend(page)
        if len(page) < _PAGE_SIZE:
            return records
        offset += _PAGE_SIZE


class DailySalesXlsxWriter:
    """Escribe el detalle de ventas y la conciliacion de pagos."""

    SALES_HEADERS = [
        "Fecha",
        "Tipo de pago",
        "Producto",
        "Referencia de producto",
        "Nombre POS",
        "Cantidad",
        "Precio",
        "Referencia de venta",
        "Sesion POS",
    ]

    def save(self, sales: Sequence[PosSale], output_path: Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sales_sheet = workbook.active
        sales_sheet.title = "Ventas PoS"
        sales_sheet.append(self.SALES_HEADERS)
        self._style_header(sales_sheet)
        for sale in sales:
            sales_sheet.append(
                [
                    _xlsx_date_value(sale.date),
                    _excel_value(sale.payment_type),
                    _excel_value(sale.product),
                    _excel_value(sale.product_reference),
                    _excel_value(sale.pos_name),
                    float(sale.quantity),
                    float(sale.price),
                    _excel_value(sale.sale_reference),
                    _excel_value(sale.pos_session),
                ]
            )
        self._format_sales_sheet(sales_sheet)

        payment_rows: list[tuple[PosSale, PosPayment]] = []
        seen_payment_ids: set[int] = set()
        payment_totals: dict[str, Decimal] = {}
        for sale in sales:
            for payment in sale.payments:
                if payment.payment_id in seen_payment_ids:
                    continue
                seen_payment_ids.add(payment.payment_id)
                payment_rows.append((sale, payment))
                payment_totals[payment.payment_type] = (
                    payment_totals.get(payment.payment_type, Decimal("0"))
                    + payment.amount
                )

        payment_sheet = workbook.create_sheet("Pagos PoS")
        payment_sheet.append(
            [
                "Fecha",
                "Referencia de venta",
                "Tipo de pago",
                "Importe pagado",
                "Nombre POS",
                "Sesion POS",
            ]
        )
        self._style_header(payment_sheet)
        for sale, payment in payment_rows:
            payment_sheet.append(
                [
                    _xlsx_date_value(sale.date),
                    _excel_value(sale.sale_reference),
                    _excel_value(payment.payment_type),
                    float(payment.amount),
                    _excel_value(sale.pos_name),
                    _excel_value(sale.pos_session),
                ]
            )
        self._format_payment_sheet(payment_sheet)

        summary_sheet = workbook.create_sheet("Resumen pagos")
        summary_sheet.append(["Tipo de pago", "Total pagado"])
        self._style_header(summary_sheet)
        for payment_type, total in payment_totals.items():
            summary_sheet.append([_excel_value(payment_type), float(total)])
        summary_sheet.column_dimensions["A"].width = 18
        summary_sheet.column_dimensions["B"].width = 16
        summary_sheet.auto_filter.ref = summary_sheet.dimensions
        for cell in summary_sheet["B"][1:]:
            cell.number_format = "0.00"

        workbook.save(output_path)
        return output_path

    @staticmethod
    def _style_header(worksheet: Any) -> None:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    @staticmethod
    def _format_sales_sheet(worksheet: Any) -> None:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        widths = {
            "A": 20,
            "B": 18,
            "C": 28,
            "D": 24,
            "E": 22,
            "F": 12,
            "G": 14,
            "H": 24,
            "I": 24,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        for cell in worksheet["A"][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
        for cell in worksheet["F"][1:]:
            cell.number_format = "0.######"
        for cell in worksheet["G"][1:]:
            cell.number_format = "0.00"

    @staticmethod
    def _format_payment_sheet(worksheet: Any) -> None:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        widths = {"A": 20, "B": 24, "C": 18, "D": 16, "E": 22, "F": 24}
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        for cell in worksheet["A"][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
        for cell in worksheet["D"][1:]:
            cell.number_format = "0.00"


def _odoo_datetime(value: datetime) -> str:
    return value.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def _parse_odoo_datetime(value: Any, target_timezone: ZoneInfo) -> datetime:
    if isinstance(value, datetime):
        parsed_value = value
    elif isinstance(value, date):
        parsed_value = datetime.combine(value, time.min)
    else:
        if not value:
            raise ValueError("La orden de Odoo no tiene fecha de venta.")
        parsed_value = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    if parsed_value.tzinfo is None:
        parsed_value = parsed_value.replace(tzinfo=timezone.utc)
    return parsed_value.astimezone(target_timezone)


def _xlsx_date_value(value: datetime) -> datetime:
    if value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _excel_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _decimal_value(value: Any) -> Decimal:
    return Decimal(str(value if value not in (None, False) else 0))


def _relation_id(value: Any) -> int | None:
    if isinstance(value, (list, tuple)):
        value = value[0] if value else None
    if value in (None, False, ""):
        return None
    return int(value)


def _relation_name(value: Any) -> str:
    if isinstance(value, (list, tuple)):
        return str(value[1]) if len(value) > 1 and value[1] else ""
    return ""


def _relation_ids(values: Iterable[Any]) -> list[int]:
    return list(
        dict.fromkeys(
            relation_id
            for value in values
            if (relation_id := _relation_id(value)) is not None
        )
    )


def _list_relation_ids(values: Iterable[Any]) -> list[int]:
    return list(
        dict.fromkeys(
            relation_id
            for value in values
            for relation_id in _relation_ids(value or [])
        )
    )


def _payments_for_order(
    payment_ids: Iterable[Any],
    payment_by_id: dict[int, PosPayment],
) -> list[PosPayment]:
    payments: list[PosPayment] = []
    for payment_id in payment_ids or []:
        payment_id = _relation_id(payment_id)
        if payment_id is None:
            continue
        payment = payment_by_id.get(payment_id)
        if payment is not None:
            payments.append(payment)
    return payments


def _payment_type_label(payments: Iterable[PosPayment]) -> str:
    payment_types: list[str] = []
    for payment in payments:
        if payment.payment_type not in payment_types:
            payment_types.append(payment.payment_type)
    return ", ".join(payment_types) or "Sin pago"


def _parse_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as error:
        raise argparse.ArgumentTypeError(
            f"Fecha invalida, usa el formato YYYY-MM-DD: {value}"
        ) from error


def _default_output_path(start_date: date, end_date: date | None) -> Path:
    suffix = end_date.isoformat() if end_date else "hasta-hoy"
    return (
        project_root()
        / "reports"
        / f"pos-sales-{start_date.isoformat()}-a-{suffix}.xlsx"
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Genera un reporte XLSX de ventas de PoS de Odoo."
    )
    parser.add_argument(
        "--start-date",
        required=True,
        type=_parse_date,
        help="Fecha inicial inclusive, en formato YYYY-MM-DD.",
    )
    parser.add_argument(
        "--end-date",
        type=_parse_date,
        help="Fecha final inclusive. Si se omite, consulta hasta la actualidad.",
    )
    parser.add_argument(
        "--output-path",
        type=Path,
        help="Ruta del XLSX de salida.",
    )
    return parser


def cli_main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    output_path = args.output_path or _default_output_path(
        args.start_date,
        args.end_date,
    )
    root = project_root()
    client = XmlRpcOdooClient(
        OdooSettings(_env_file=root / ".env"),
        access_mode=AccessMode.READ_ONLY,
    )
    config = DailySalesReportConfig(
        start_date=args.start_date,
        end_date=args.end_date,
        output_path=output_path,
    )
    generated_path = generate_daily_sales_report(client, config)
    print(f"Reporte generado en: {generated_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(cli_main())
