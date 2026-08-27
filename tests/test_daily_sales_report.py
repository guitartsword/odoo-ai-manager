from collections.abc import Mapping, Sequence
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from odoo_ai_manager.modules.pos.daily_sales_report import (
    DailySalesReportConfig,
    DailySalesXlsxWriter,
    generate_daily_sales_report,
)
from odoo_ai_manager.modules.pos.entities import PosSale


class FakeOdooReadClient:
    access_mode = "read_only"
    user_id = 7
    user_timezone = ZoneInfo("America/Mexico_City")
    server_version = "16.0"
    company_id = 42
    company_name = "Bee Lovely"

    def __init__(self) -> None:
        self.calls: list[dict[str, Any]] = []

    def search_read(
        self,
        model: str,
        domain: Sequence[Sequence[Any]],
        fields: Sequence[str],
        *,
        offset: int = 0,
        limit: int = 1000,
        order: str | None = None,
        context: Mapping[str, Any] | None = None,
    ) -> list[dict[str, Any]]:
        self.calls.append(
            {
                "model": model,
                "domain": domain,
                "fields": fields,
                "offset": offset,
                "limit": limit,
            }
        )
        if model == "pos.order.line":
            return [
                {
                    "id": 10,
                    "order_id": [100, "Order 100"],
                    "product_id": [1, "Cafe"],
                    "qty": 1,
                    "price_unit": 1695,
                }
            ]
        if model == "pos.order":
            return [
                {
                    "id": 100,
                    "name": "POS/2026/0001",
                    "pos_reference": "Order 100",
                    "date_order": "2026-07-01 06:48:59",
                    "config_id": [20, "Bee Lovely"],
                    "session_id": [40, "POS Session 2026/07/01"],
                    "payment_ids": [500, 501],
                }
            ]
        if model == "product.product":
            return [{"id": 1, "default_code": "BEE-0218"}]
        if model == "pos.payment":
            return [
                {
                    "id": 500,
                    "payment_method_id": [30, "Visa"],
                    "amount": 1000,
                },
                {
                    "id": 501,
                    "payment_method_id": [31, "Efectivo"],
                    "amount": 695,
                },
            ]
        raise AssertionError(f"Unexpected model: {model}")


def test_daily_sales_report_uses_user_timezone_and_writes_payment_breakdown(
    tmp_path: Path,
) -> None:
    client = FakeOdooReadClient()
    output_path = tmp_path / "reports" / "sales.xlsx"
    config = DailySalesReportConfig(
        start_date=datetime(2026, 7, 1).date(),
        end_date=datetime(2026, 7, 1).date(),
        output_path=output_path,
    )

    result = generate_daily_sales_report(client, config)

    assert result == output_path
    line_call = next(call for call in client.calls if call["model"] == "pos.order.line")
    assert ("order_id.date_order", ">=", "2026-07-01 06:00:00") in line_call["domain"]
    assert ("order_id.date_order", "<", "2026-07-02 06:00:00") in line_call["domain"]
    assert ("order_id.company_id", "=", 42) in line_call["domain"]

    workbook = load_workbook(output_path)
    sales_sheet = workbook["Ventas PoS"]
    assert sales_sheet["A2"].value == datetime(2026, 7, 1, 0, 48, 59)
    assert sales_sheet["H2"].value == "Order 100"
    assert sales_sheet["I2"].value == "POS Session 2026/07/01"

    payments_sheet = workbook["Pagos PoS"]
    assert [list(row) for row in payments_sheet.values][1:] == [
        [
            datetime(2026, 7, 1, 0, 48, 59),
            "Order 100",
            "Visa",
            1000,
            "Bee Lovely",
            "POS Session 2026/07/01",
        ],
        [
            datetime(2026, 7, 1, 0, 48, 59),
            "Order 100",
            "Efectivo",
            695,
            "Bee Lovely",
            "POS Session 2026/07/01",
        ],
    ]

    summary_sheet = workbook["Resumen pagos"]
    assert [list(row) for row in summary_sheet.values][1:] == [
        ["Visa", 1000],
        ["Efectivo", 695],
    ]


def test_daily_sales_xlsx_escapes_formula_like_text(tmp_path: Path) -> None:
    output_path = tmp_path / "sales.xlsx"
    sale = PosSale(
        date=datetime(2026, 7, 1, 0, 0),
        payment_type="Efectivo",
        product="=HYPERLINK(\"https://example.com\")",
        product_reference="@reference",
        pos_name="Tienda",
        sale_reference="-sale",
        pos_session="+session",
        quantity=Decimal("1"),
        price=Decimal("10"),
    )

    DailySalesXlsxWriter().save([sale], output_path)

    sheet = load_workbook(output_path)["Ventas PoS"]
    assert sheet["C2"].value == "'=HYPERLINK(\"https://example.com\")"
    assert sheet["D2"].value == "'@reference"
    assert sheet["H2"].value == "'-sale"
    assert sheet["I2"].value == "'+session"
